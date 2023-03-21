import os
import pandas as pd
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
from PyQt5.uic import loadUi
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import warnings


pd.set_option("display.max_rows", 1000)
pd.set_option("display.max_columns", 100)
pd.set_option("display.precision", 100)

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
current_dir=os.path.dirname(os.path.abspath("__file__"))
df = pd.read_excel(os.path.join(current_dir, 'Регламент_инструкции.xlsx'), "Регламент")

class MyWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        # добавляем виджеты на форму
        self.date_label = QtWidgets.QLabel('Дата:')
        self.date_edit = QtWidgets.QLineEdit()
        self.hours_label = QtWidgets.QLabel('Машиночасы:')
        self.hours_edit = QtWidgets.QLineEdit()
        self.mark_label = QtWidgets.QLabel('Марка техники:')
        self.mark_edit = QtWidgets.QLineEdit()
        self.inventory_label = QtWidgets.QLabel('Инвентарный номер:')
        self.inventory_edit = QtWidgets.QLineEdit()
        self.to_label = QtWidgets.QLabel('Номер ТО:')
        self.to_edit = QtWidgets.QLineEdit()
        self.submit_button = QtWidgets.QPushButton('Отправить')

        # добавляем таблицу для вывода результатов
        self.result_table = QtWidgets.QTableWidget()
        self.result_table.setColumnCount(3)
        self.result_table.setHorizontalHeaderLabels(['Работы', 'Количество план', 'Количество факт'])

        # размещаем виджеты на форме
        form_layout = QtWidgets.QFormLayout()
        form_layout.addRow(self.date_label, self.date_edit)
        form_layout.addRow(self.hours_label, self.hours_edit)
        form_layout.addRow(self.mark_label, self.mark_edit)
        form_layout.addRow(self.inventory_label, self.inventory_edit)
        form_layout.addRow(self.to_label, self.to_edit)
        form_layout.addWidget(self.submit_button)

        main_layout = QtWidgets.QVBoxLayout()
        main_layout.addLayout(form_layout)
        main_layout.addWidget(self.result_table)

        self.setLayout(main_layout)

        # связываем сигнал нажатия кнопки с методом on_submit_clicked
        self.submit_button.clicked.connect(self.on_submit_clicked)

        self.save_button = QtWidgets.QPushButton('Сохранить в Excel')
        form_layout.addRow(self.save_button)
        self.save_button.clicked.connect(self.save_to_excel)

    def on_submit_clicked(self):
        mark = self.mark_edit.text()
        to_number = self.to_edit.text()

        # поиск данных в датафрейме
        data = df[(df['Марка техники'] == mark) & (df['Вид ТО'] == to_number)]

        if not data.empty:
            # задаем данные для таблицы
            works = data['Наименование'].tolist()
            amounts = data['Кол-во'].tolist()

            self.result_table.setRowCount(len(works))

            for i, (work, amount) in enumerate(zip(works, amounts)):
                self.result_table.setItem(i, 0, QtWidgets.QTableWidgetItem(work))
                self.result_table.setItem(i, 1, QtWidgets.QTableWidgetItem(str(amount)))

                edit = QtWidgets.QLineEdit()
                self.result_table.setCellWidget(i, 2, edit)
        else:
            QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Неверный номер ТО для выбранной марки техники "{mark}"')

    def save_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, 'Сохранить файл', '', 'Excel files (*.xlsx)')

        if file_path:
            wb = Workbook()
            ws = wb.active

            # Записываем заголовки таблицы в первую строку
            for i in range(self.result_table.columnCount()):
                cell = ws.cell(row=1, column=i + 1)
                cell.value = self.result_table.horizontalHeaderItem(i).text()

            # Записываем данные из таблицы в ячейки
            for row in range(self.result_table.rowCount()):
                for col in range(self.result_table.columnCount()):
                    cell = ws.cell(row=row + 2, column=col + 1)
                    item = self.result_table.item(row, col)
                    if item is not None:
                        cell.value = item.text()
                    else:
                        cell.value = self.result_table.cellWidget(row, col).text()

            wb.save(file_path)


if __name__ == '__main__':
    app = QtWidgets.QApplication([])
    window = MyWindow()
    window.show()
    app.exec_()